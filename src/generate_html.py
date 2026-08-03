"""
Turns data/job_tracker.xlsx into a plain, static HTML page at
docs/index.html, so it can be published via GitHub Pages — one link,
opens in any browser, no Excel and no download needed.

Renders both sheets as clearly separated sections, in order:
  1. "Jobs" (Munich, Zurich, Singapore, Basel, Bern, Geneva, Lausanne,
     Lucerne — green highlight for new rows)
  2. "Dream Cities" (blue highlight for new rows)

Called automatically at the end of both src/main.py (daily scrape) and
src/reset_tracker.py (monthly reset), so the page is always in sync
with whatever's actually in the Excel file. Not meant to be run on its
own, though `python -m src.generate_html` works fine for testing.

(There used to be a third "Global Top Picks" wildcard section here —
removed per request, since Singapore and the Swiss cities it partly
existed to surface are now proper main-list cities instead.)

PRIVACY NOTE: GitHub Pages sites are public to anyone with the link,
even when the repository itself is private (unless you're on GitHub
Enterprise). This page will show your company/job data to anyone who
finds the URL — see SETUP_GUIDE.md, "Enabling the public webpage view."
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import yaml
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.tracker import (
    COLUMNS, NEW_ROW_FILL,
    DREAM_SHEET_NAME, DREAM_COLUMNS, DREAM_NEW_ROW_FILL,
)

ROOT = Path(__file__).resolve().parent.parent
TRACKER_FILE = ROOT / "data" / "job_tracker.xlsx"
OUTPUT_FILE = ROOT / "docs" / "index.html"
CV_PROFILE_FILE = ROOT / "config" / "cv_profile.yaml"


def _rgb_matches(cell, fill) -> bool:
    cell_rgb = getattr(cell.fill.start_color, "rgb", None)
    fill_rgb = getattr(fill.start_color, "rgb", None)
    if not isinstance(cell_rgb, str) or not isinstance(fill_rgb, str):
        return False
    return cell_rgb[-6:] == fill_rgb[-6:]  # compare ignoring alpha prefix


def _escape(value) -> str:
    if value is None:
        return ""
    text = str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _render_table(
    ws: Worksheet | None, columns: list[str], new_row_fill, empty_message: str, row_class: str,
) -> tuple[str, int]:
    """Returns (rows_html, total_row_count) for one sheet."""
    if ws is None or ws.max_row < 2:
        return f"<tr><td colspan='{len(columns)}'>{_escape(empty_message)}</td></tr>", 0

    url_col = columns.index("URL") + 1
    title_col = columns.index("Job Title")

    rows_html_parts = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, len(columns) + 1)]
        if all(v is None for v in values):
            continue  # skip fully-empty rows
        is_highlighted = _rgb_matches(ws.cell(row=row, column=1), new_row_fill)
        url = values[url_col - 1] or "#"

        cells = []
        for i, val in enumerate(values):
            if i == title_col and url and url != "#":
                cells.append(f"<td><a href='{_escape(url)}' target='_blank' rel='noopener'>{_escape(val)}</a></td>")
            elif i == url_col - 1:
                continue  # URL folded into the Job Title link, not shown as its own column
            else:
                cells.append(f"<td>{_escape(val)}</td>")
        cls = f" class='{row_class}'" if is_highlighted else ""
        rows_html_parts.append(f"<tr{cls}>{''.join(cells)}</tr>")

    rows_html = "\n".join(rows_html_parts) if rows_html_parts else f"<tr><td colspan='{len(columns)}'>{_escape(empty_message)}</td></tr>"
    return rows_html, len(rows_html_parts)


def _section_html(title: str, columns: list[str], rows_html: str, total: int, legend_color: str, legend_label: str) -> str:
    display_columns = [c for c in columns if c != "URL"]
    header_html = "".join(f"<th>{_escape(c)}</th>" for c in display_columns)
    return f"""
<h2>{_escape(title)}</h2>
<div class="meta">
  {total} tracked postings &middot;
  <span class="legend" style="background:{legend_color};"></span>{_escape(legend_label)}
</div>
<div class="scroll-wrap">
<table>
<thead><tr>{header_html}</tr></thead>
<tbody>
{rows_html}
</tbody>
</table>
</div>
"""


def generate(tracker_path: Path = TRACKER_FILE, output_path: Path = OUTPUT_FILE) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    cv_profile = {}
    if CV_PROFILE_FILE.exists():
        with open(CV_PROFILE_FILE, "r", encoding="utf-8") as f:
            cv_profile = yaml.safe_load(f) or {}
    main_min_score = cv_profile.get("main_min_score", 0)
    dream_min_score = cv_profile.get("dream_city_min_score", 0)

    jobs_ws = None
    dream_ws = None
    if tracker_path.exists():
        wb = load_workbook(tracker_path)
        jobs_ws = wb["Jobs"] if "Jobs" in wb.sheetnames else wb.active
        dream_ws = wb[DREAM_SHEET_NAME] if DREAM_SHEET_NAME in wb.sheetnames else None

    jobs_rows_html, jobs_total = _render_table(
        jobs_ws, COLUMNS, NEW_ROW_FILL, "No data yet — the scraper hasn't run.", "new-row"
    )
    dream_rows_html, dream_total = _render_table(
        dream_ws, DREAM_COLUMNS, DREAM_NEW_ROW_FILL, "No matches yet in the dream-city list.", "dream-new-row"
    )

    jobs_title = "Munich, Zurich, Singapore & Swiss Cities" + (f" (score {main_min_score}+ only)" if main_min_score else "")
    dream_title = "Dream Cities" + (f" (score {dream_min_score}+ only)" if dream_min_score else "")

    jobs_section = _section_html(
        jobs_title, COLUMNS, jobs_rows_html, jobs_total, "#D1FAE5", "added since your last check"
    )
    dream_section = _section_html(
        dream_title, DREAM_COLUMNS, dream_rows_html, dream_total, "#DBEAFE", "added since your last check"
    )

    updated = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Job Tracker</title>
<style>
  body {{ font-family: -apple-system, Segoe UI, Roboto, Arial, sans-serif; margin: 0; padding: 16px; background: #f9fafb; color: #111827; }}
  h1 {{ font-size: 1.3rem; margin-bottom: 4px; }}
  h2 {{ font-size: 1.05rem; margin: 28px 0 4px 0; }}
  .page-meta {{ color: #6b7280; font-size: 0.85rem; margin-bottom: 8px; }}
  .meta {{ color: #6b7280; font-size: 0.85rem; margin-bottom: 10px; }}
  .legend {{ display: inline-block; width: 12px; height: 12px; border: 1px solid rgba(0,0,0,0.15); margin-right: 6px; vertical-align: middle; }}
  table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
  th, td {{ padding: 8px 10px; text-align: left; border-bottom: 1px solid #e5e7eb; font-size: 0.85rem; vertical-align: top; }}
  th {{ background: #1F2937; color: white; position: sticky; top: 0; white-space: nowrap; }}
  tr.new-row {{ background: #D1FAE5; }}
  tr.dream-new-row {{ background: #DBEAFE; }}
  tr:hover {{ background: #f3f4f6; }}
  tr.new-row:hover {{ background: #bbf7d0; }}
  tr.dream-new-row:hover {{ background: #bfdbfe; }}
  a {{ color: #2563eb; text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .scroll-wrap {{ overflow-x: auto; margin-bottom: 8px; }}
  hr {{ border: none; border-top: 1px solid #e5e7eb; margin: 24px 0; }}
  @media (max-width: 700px) {{
    th, td {{ font-size: 0.75rem; padding: 6px; }}
    body {{ padding: 8px; }}
  }}
</style>
</head>
<body>
<h1>Job Tracker</h1>
<div class="page-meta">last updated {updated}</div>
{jobs_section}
<hr>
{dream_section}
</body>
</html>
"""
    output_path.write_text(html, encoding="utf-8")


if __name__ == "__main__":
    generate()
