"""
Manages the persistent Excel tracker file, which holds THREE sheets,
in this order (top to bottom, matching the requested reading order):

  1. "Jobs" — Munich and Zurich only. The original core list.
  2. "Singapore & Swiss Cities" — Singapore, Basel, Bern, Geneva,
     Lausanne, Lucerne. Promoted from Dream Cities to full main-list
     status, but kept as its own visually separate section rather
     than merged into "Jobs" — otherwise Munich/Zurich's much higher
     posting volume buries these cities' (genuinely fewer) matches at
     the bottom of one score-sorted list, making it look like nothing
     is there even when something is.
  3. "Dream Cities" — the remaining aspirational city list
     (Copenhagen, Oslo, Helsinki, Vienna, Berlin, Amsterdam,
     Rotterdam, Vancouver, Perth, Melbourne, Sydney).

All three sheets share the same behaviour:
  - Each run, only genuinely NEW postings (by URL) are added.
  - Postings already in a sheet are left alone (deduped by URL,
    per-sheet — a URL only needs to be unique within its own sheet).
  - Every run, rows scoring below that sheet's min_score are pruned —
    this applies to rows already sitting in the sheet too, not just
    new ones, so lowering/raising the threshold in cv_profile.yaml
    actually cleans up the sheet, not just gates future additions.
  - Each sheet is re-sorted by Relevance Score (highest first) every
    run, so the best matches are always at the top.
  - archive_and_reset() (called from reset_tracker.py) archives ALL
    THREE sheets together as one file and starts fresh — same monthly
    cadence, same manual-trigger option, for all of them.

Column history for the "Jobs" sheet (kept here so future-you
understands old files):
  v1: First Seen, Last Seen, Company, City, Job Title, Relevance Score,
      German Required, Location, URL
  v2: added Job Posted; renamed score column to "Relevance Score (1-10)"
  v3: added Location Confirmed
  v4: dropped First Seen and Last Seen; sheet sorted by relevance
      instead of being append-only
  v5: added the separate "Dream Cities" sheet
  v6: dropped "Location Confirmed" — only confirmed-location jobs are
      kept at all now (matcher.py), so the column was always "Yes"
      and stopped being useful information.
  v7: a "Global Top Picks" wildcard sheet existed briefly between v6
      and now — removed entirely per request.
  v8 (current): Singapore + the promoted Swiss cities briefly lived
      inside "Jobs" alongside Munich/Zurich — split back out into
      their own "Singapore & Swiss Cities" sheet, per request, so
      they're not buried by Munich/Zurich's higher volume.
Migration reads whatever columns an existing sheet has by NAME, not
position, and rebuilds it against the current column list — so it
doesn't matter which older version your file is on. If your file
still has an old "Global Top Picks" or "_wildcard_history" sheet from
before v7, load_or_create just leaves it alone (inert leftover, not
migrated) — delete it by hand in Excel if you want it gone.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

COLUMNS = [
    "Job Posted",
    "Company",
    "City",
    "Job Title",
    "Relevance Score (1-10)",
    "Location",
    "URL",
]
COLUMN_WIDTHS = [14, 22, 9, 40, 12, 30, 60]

# Singapore & Swiss Cities shares the exact same schema as Jobs (no
# Country column needed — same reasoning as Jobs: obvious from City).
SWISS_SG_SHEET_NAME = "Singapore & Swiss Cities"
SWISS_SG_COLUMNS = COLUMNS
SWISS_SG_COLUMN_WIDTHS = COLUMN_WIDTHS

DREAM_SHEET_NAME = "Dream Cities"
DREAM_COLUMNS = [
    "Job Posted",
    "Company",
    "City",
    "Country",
    "Job Title",
    "Relevance Score (1-10)",
    "Location",
    "URL",
]
DREAM_COLUMN_WIDTHS = [14, 22, 11, 12, 40, 12, 30, 60]

# Old header names that should map onto a current column, so a rename
# (like "Relevance Score" -> "Relevance Score (1-10)") doesn't strand
# existing data. Names not listed here and not in the current column
# list are dropped (that's how First Seen / Last Seen / Location
# Confirmed disappear on migration).
HEADER_ALIASES = {
    "Relevance Score": "Relevance Score (1-10)",
}

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NEW_ROW_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # green
SWISS_SG_NEW_ROW_FILL = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")  # amber
DREAM_NEW_ROW_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # blue


def _style_header(ws: Worksheet, columns: list[str], widths: list[int]) -> None:
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _new_workbook() -> Workbook:
    """Creates a fresh workbook with ALL THREE sheets, correctly
    headered, in the requested order: Jobs -> Singapore & Swiss
    Cities -> Dream Cities."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append(COLUMNS)
    _style_header(ws, COLUMNS, COLUMN_WIDTHS)

    swiss_sg_ws = wb.create_sheet(SWISS_SG_SHEET_NAME)
    swiss_sg_ws.append(SWISS_SG_COLUMNS)
    _style_header(swiss_sg_ws, SWISS_SG_COLUMNS, SWISS_SG_COLUMN_WIDTHS)

    dream_ws = wb.create_sheet(DREAM_SHEET_NAME)
    dream_ws.append(DREAM_COLUMNS)
    _style_header(dream_ws, DREAM_COLUMNS, DREAM_COLUMN_WIDTHS)
    return wb


def _migrate_sheet(wb: Workbook, sheet_name: str, columns: list[str], widths: list[int], sheet_index: int) -> None:
    """
    Generic migration for any of the three sheets: read every row
    into a dict keyed by (aliased) old header name, drop any row that
    was "Unconfirmed" under the old Location Confirmed scheme, then
    rebuild the sheet from scratch against the current column list.
    """
    ws = wb[sheet_name]
    existing_headers = [c.value for c in ws[1]]
    if existing_headers == columns:
        return  # already current

    mapped_headers = [HEADER_ALIASES.get(h, h) for h in existing_headers]
    location_confirmed_idx = None
    if "Location Confirmed" in existing_headers:
        location_confirmed_idx = existing_headers.index("Location Confirmed")

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        if location_confirmed_idx is not None:
            confirmed_val = ws.cell(row=row_idx, column=location_confirmed_idx + 1).value
            if confirmed_val != "Yes":
                continue  # drop rows that weren't a confirmed-location match
        row_dict = {}
        for col_idx, header in enumerate(mapped_headers, start=1):
            if header in columns:
                row_dict[header] = ws.cell(row=row_idx, column=col_idx).value
        if row_dict:
            rows.append(row_dict)

    wb.remove(ws)
    new_ws = wb.create_sheet(sheet_name, sheet_index)
    new_ws.append(columns)
    for row_dict in rows:
        new_ws.append([row_dict.get(col, "") for col in columns])
    _style_header(new_ws, columns, widths)


def _ensure_sheet(wb: Workbook, sheet_name: str, columns: list[str], widths: list[int], sheet_index: int) -> None:
    """Creates the sheet fresh if missing, or migrates it in place if
    it already exists under an older schema."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name, sheet_index)
        ws.append(columns)
        _style_header(ws, columns, widths)
    else:
        _migrate_sheet(wb, sheet_name, columns, widths, sheet_index)


def load_or_create(path: Path) -> Workbook:
    if not path.exists():
        return _new_workbook()

    wb = load_workbook(path)
    _migrate_sheet(wb, "Jobs", COLUMNS, COLUMN_WIDTHS, 0)
    _ensure_sheet(wb, SWISS_SG_SHEET_NAME, SWISS_SG_COLUMNS, SWISS_SG_COLUMN_WIDTHS, 1)
    _ensure_sheet(wb, DREAM_SHEET_NAME, DREAM_COLUMNS, DREAM_COLUMN_WIDTHS, 2)

    # Deliberately NOT touching "Global Top Picks" or "_wildcard_history"
    # if an old file still has them — that feature was removed; this
    # just leaves those sheets as inert leftovers rather than deleting
    # or migrating them. Delete the tabs by hand in Excel if you want
    # them gone from an existing file.

    return wb


def _existing_urls(ws: Worksheet, columns: list[str]) -> set[str]:
    url_col = columns.index("URL") + 1
    return {
        ws.cell(row=row, column=url_col).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=url_col).value
    }


def _prune_below_score(ws: Worksheet, columns: list[str], min_score: int) -> int:
    """Deletes any row scoring below min_score. Runs every update, so
    it retroactively cleans rows already in the sheet too, not just
    new ones going forward. Returns how many rows were removed."""
    if not min_score:
        return 0
    score_col = columns.index("Relevance Score (1-10)") + 1
    rows_to_delete = [
        row
        for row in range(2, ws.max_row + 1)
        if not (
            isinstance(ws.cell(row=row, column=score_col).value, (int, float))
            and ws.cell(row=row, column=score_col).value >= min_score
        )
    ]
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
    return len(rows_to_delete)


def _sort_by_relevance(
    ws: Worksheet, columns: list[str], fill: PatternFill, newly_added_urls: set[str]
) -> None:
    """Re-sorts all data rows by Relevance Score (highest first), then
    re-applies the "new" highlight to whichever URLs were added this
    run (their row position may have moved during the sort)."""
    score_col = columns.index("Relevance Score (1-10)")
    url_col = columns.index("URL")

    rows = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row=row, column=c).value for c in range(1, len(columns) + 1)]
        rows.append(values)

    rows.sort(key=lambda v: (v[score_col] if isinstance(v[score_col], (int, float)) else 0), reverse=True)

    for row in range(2, ws.max_row + 1):
        for col in range(1, len(columns) + 1):
            ws.cell(row=row, column=col).value = None
            ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

    for i, values in enumerate(rows):
        row_idx = i + 2
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
        if values[url_col] in newly_added_urls:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill


def _update_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[str],
    fill: PatternFill,
    new_jobs: list[dict[str, Any]],
    row_builder: Callable[[dict[str, Any], str], list],
    min_score: int = 0,
) -> dict[str, int]:
    ws = wb[sheet_name]
    existing_urls = _existing_urls(ws, columns)

    added = 0
    already_tracked = 0
    newly_added_urls = set()
    for job in new_jobs:
        url = job.get("url", "")
        if not url:
            continue
        if url in existing_urls:
            already_tracked += 1
            continue
        ws.append(row_builder(job, url))
        newly_added_urls.add(url)
        existing_urls.add(url)
        added += 1

    pruned = _prune_below_score(ws, columns, min_score)
    _sort_by_relevance(ws, columns, fill, newly_added_urls)
    return {"added": added, "already_tracked": already_tracked, "pruned": pruned, "total_rows": ws.max_row - 1}


def _build_jobs_row(job: dict[str, Any], url: str) -> list:
    return [
        job.get("posted_date", ""),
        job.get("company", ""),
        job.get("city", ""),
        job.get("title", ""),
        job.get("relevance_score", 1),
        job.get("location", ""),
        url,
    ]


def _build_dream_row(job: dict[str, Any], url: str) -> list:
    return [
        job.get("posted_date", ""),
        job.get("company", ""),
        job.get("city", ""),
        job.get("country", ""),
        job.get("title", ""),
        job.get("relevance_score", 1),
        job.get("location", ""),
        url,
    ]


def update_tracker(path: Path, new_jobs: list[dict[str, Any]], min_score: int = 0) -> dict[str, int]:
    """Updates the "Jobs" sheet (Munich and Zurich only). Saves the file."""
    wb = load_or_create(path)
    summary = _update_sheet(wb, "Jobs", COLUMNS, NEW_ROW_FILL, new_jobs, _build_jobs_row, min_score=min_score)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summary


def update_swiss_sg_tracker(path: Path, new_jobs: list[dict[str, Any]], min_score: int = 0) -> dict[str, int]:
    """Updates the "Singapore & Swiss Cities" sheet. Saves the file."""
    wb = load_or_create(path)
    summary = _update_sheet(
        wb, SWISS_SG_SHEET_NAME, SWISS_SG_COLUMNS, SWISS_SG_NEW_ROW_FILL, new_jobs, _build_jobs_row, min_score=min_score
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summary


def update_dream_tracker(path: Path, new_jobs: list[dict[str, Any]], min_score: int = 0) -> dict[str, int]:
    """Updates the "Dream Cities" sheet. Saves the file. Call each of
    the three update_* functions separately (as main.py does) — each
    opens, updates its own sheet, and saves; the file just gets saved
    three times per run, which is harmless."""
    wb = load_or_create(path)
    summary = _update_sheet(
        wb, DREAM_SHEET_NAME, DREAM_COLUMNS, DREAM_NEW_ROW_FILL, new_jobs, _build_dream_row, min_score=min_score
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return summary


def archive_and_reset(path: Path, archive_dir: Path) -> Path | None:
    """
    Moves the current tracker (all three sheets) to
    archive/job_tracker_YYYY-MM.xlsx and creates a fresh empty tracker
    (all three sheets, empty) at `path`. Returns the archive path, or
    None if there was nothing to archive.
    """
    if not path.exists():
        _new_workbook().save(path)
        return None

    archive_dir.mkdir(parents=True, exist_ok=True)
    month_tag = dt.date.today().strftime("%Y-%m")
    archive_path = archive_dir / f"job_tracker_{month_tag}.xlsx"

    counter = 2
    final_path = archive_path
    while final_path.exists():
        final_path = archive_dir / f"job_tracker_{month_tag}_v{counter}.xlsx"
        counter += 1

    path.rename(final_path)
    _new_workbook().save(path)
    return final_path
